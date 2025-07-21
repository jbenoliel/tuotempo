#!/usr/bin/env python3
"""
Corrección específica para registros "No Interesado" con subestados faltantes
"""
import os
import json
import logging
import requests
from openpyxl import load_workbook
from dotenv import load_dotenv
from mapeo_inteligente_segurcaixa import MapeadorInteligente

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class CorrectorNoInteresados:
    def __init__(self):
        load_dotenv()
        self.api_base_url = "https://tuotempo-apis-production.up.railway.app"
        self.mapeador = MapeadorInteligente()
        self.corregidos = 0
        
    def extract_json_fields(self, collected_info_value):
        """Extraer campos del JSON collectedInfo"""
        fields = {
            "firstName": "", "lastName": "", "phoneNumber": "", "nombreClinica": "",
            "noInteresado": False, "volverALlamar": False, "conPack": False, "sinPack": False,
            "razonNoInteres": ""
        }
        
        if not collected_info_value:
            return fields
        
        try:
            if isinstance(collected_info_value, str):
                data = json.loads(collected_info_value)
            elif isinstance(collected_info_value, dict):
                data = collected_info_value
            else:
                return fields
            
            for field in fields.keys():
                if field in data:
                    fields[field] = data[field]
        except:
            pass
        
        return fields
    
    def corregir_registro(self, nombre, telefono, collected_info):
        """Corregir un registro No Interesado específico"""
        
        # Payload simplificado para No Interesado
        payload = {
            "telefono": telefono,
            "noInteresado": True,
            "razonNoInteres": "Cliente no interesado",  # Texto simple
            "codigoNoInteres": "otros"
        }
        
        logger.info(f"Corrigiendo {nombre} ({telefono})")
        logger.info(f"Payload: {payload}")
        
        try:
            url = f"{self.api_base_url}/api/actualizar_resultado"
            headers = {'Content-Type': 'application/json'}
            response = requests.post(url, json=payload, headers=headers, timeout=30)
            
            if response.status_code == 200:
                logger.info(f"✓ CORREGIDO: {nombre}")
                self.corregidos += 1
                return True
            else:
                logger.error(f"✗ ERROR {nombre}: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            logger.error(f"✗ EXCEPCION {nombre}: {e}")
            return False
    
    def procesar_archivo(self, file_path):
        """Procesar solo registros No Interesado"""
        logger.info("=== CORRECTOR DE NO INTERESADOS ===")
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Encontrar columna CollectedInfo
        collected_info_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "collectedinfo" in str(cell_value).lower():
                collected_info_col = col
                break
        
        if not collected_info_col:
            logger.error("No se encontró columna CollectedInfo")
            return
        
        # Procesar solo registros No Interesado
        registros_no_interesados = []
        
        for row in range(2, ws.max_row + 1):
            collected_info_raw = ws.cell(row=row, column=collected_info_col).value
            collected_info = self.extract_json_fields(collected_info_raw)
            
            # Solo procesar si es No Interesado
            if collected_info.get('noInteresado', False):
                nombre = f"{collected_info.get('firstName', '')} {collected_info.get('lastName', '')}".strip()
                telefono = collected_info.get('phoneNumber', '')
                
                # Limpiar teléfono
                if telefono.startswith('+34'):
                    telefono = telefono[3:]
                elif telefono.startswith('34') and len(telefono) == 11:
                    telefono = telefono[2:]
                
                registros_no_interesados.append({
                    'nombre': nombre,
                    'telefono': telefono,
                    'collected_info': collected_info
                })
        
        logger.info(f"Encontrados {len(registros_no_interesados)} registros No Interesado")
        
        # Procesar cada uno
        for registro in registros_no_interesados:
            self.corregir_registro(
                registro['nombre'], 
                registro['telefono'], 
                registro['collected_info']
            )
        
        logger.info(f"=== CORRECCIÓN COMPLETADA: {self.corregidos}/{len(registros_no_interesados)} corregidos ===")

def main():
    file_path = "C:\\Users\\jbeno\\CascadeProjects\\PearlInfo\\SegurcaixaCalls_julio18_procesar_grabaciones.xlsx"
    
    corrector = CorrectorNoInteresados()
    corrector.procesar_archivo(file_path)

if __name__ == "__main__":
    main()