#!/usr/bin/env python3
"""
Script para agregar campos status_level1 y status_level2 desde la base de datos de Railway
al archivo Excel con URLs de grabación, basándose en el número de teléfono.

Considera que:
- Los teléfonos en el Excel pueden tener +34 o 34 delante
- Los teléfonos en la BD están sin el código de país
- Normaliza los teléfonos para hacer la búsqueda correcta
"""

import os
import logging
import mysql.connector
from openpyxl import load_workbook
from dotenv import load_dotenv
from config import settings

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('agregar_status_bd.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class StatusUpdater:
    def __init__(self):
        """Inicializar el actualizador de status"""
        load_dotenv()
        self.estadisticas = {
            'total_filas': 0,
            'encontrados_bd': 0, 
            'no_encontrados': 0,
            'errores_bd': 0
        }
        
    def get_db_connection(self):
        """Establece conexión con la base de datos MySQL usando EXACTAMENTE la misma configuración que procesar_llamadas_segurcaixa.py"""
        try:
            # Configuración EXACTAMENTE igual a procesar_llamadas_segurcaixa.py
            DB_CONFIG = {
                'host': settings.DB_HOST,
                'port': settings.DB_PORT,
                'user': settings.DB_USER,
                'password': settings.DB_PASSWORD,
                'database': settings.DB_DATABASE,
                'ssl_disabled': True,  # Deshabilitar SSL para evitar errores de conexión
                'autocommit': True,
                'charset': 'utf8mb4',
                'use_unicode': True
            }
            
            logger.info(f"Conectando a Railway MySQL: {settings.DB_HOST}:{settings.DB_PORT}")
            connection = mysql.connector.connect(**DB_CONFIG)
            logger.info("Conexión a Railway MySQL establecida")
            return connection
            
        except mysql.connector.Error as err:
            logger.error(f"Error conectando a MySQL: {err}")
            # Si falla con SSL, intentar sin SSL (mismo fallback que procesar_llamadas_segurcaixa.py)
            if 'SSL' in str(err) or '2026' in str(err):
                try:
                    logger.info("Intentando conexión sin SSL...")
                    config_no_ssl = DB_CONFIG.copy()
                    config_no_ssl['ssl_disabled'] = True
                    connection = mysql.connector.connect(**config_no_ssl)
                    logger.info("Conexión sin SSL establecida")
                    return connection
                except mysql.connector.Error as err2:
                    logger.error(f"Error conectando sin SSL: {err2}")
            return None
    
    def normalize_phone(self, phone):
        """
        Normalizar número de teléfono eliminando código de país
        
        Args:
            phone (str): Número de teléfono con o sin código de país
            
        Returns:
            str: Número normalizado sin código de país (9 dígitos)
        """
        if not phone:
            return ""
        
        # Convertir a string y limpiar
        phone_str = str(phone).strip()
        
        # Eliminar caracteres no numéricos excepto el +
        phone_clean = ''.join(c for c in phone_str if c.isdigit() or c == '+')
        
        # Eliminar el + si está presente
        if phone_clean.startswith('+'):
            phone_clean = phone_clean[1:]
        
        # Si empieza con 34 y tiene 11 dígitos, quitar el 34
        if phone_clean.startswith('34') and len(phone_clean) == 11:
            phone_clean = phone_clean[2:]
        
        # Devolver solo si tiene 9 dígitos (formato español sin código de país)
        if len(phone_clean) == 9 and phone_clean.isdigit():
            return phone_clean
        
        return ""
    
    def extract_phone_from_json(self, collected_info_raw):
        """
        Extraer número de teléfono del JSON CollectedInfo
        
        Args:
            collected_info_raw: Valor crudo del JSON CollectedInfo
            
        Returns:
            str: Número de teléfono extraído o vacío
        """
        if not collected_info_raw:
            return ""
        
        try:
            import json
            if isinstance(collected_info_raw, str):
                data = json.loads(collected_info_raw)
            elif isinstance(collected_info_raw, dict):
                data = collected_info_raw
            else:
                return ""
            
            # Buscar el campo phoneNumber en el JSON
            phone = data.get('phoneNumber', '')
            return phone
            
        except (json.JSONDecodeError, TypeError) as e:
            logger.error(f"Error al decodificar JSON CollectedInfo: {e}")
            return ""
    
    def get_status_from_db(self, phone):
        """
        Obtener status_level1 y status_level2 de la base de datos
        
        Args:
            phone (str): Número de teléfono normalizado (9 dígitos)
            
        Returns:
            dict: {'status_level1': str, 'status_level2': str, 'found': bool}
        """
        if not phone:
            return {'status_level1': '', 'status_level2': '', 'found': False}
        
        # Usar conexión compartida para mejorar rendimiento
        if not hasattr(self, '_db_connection') or not self._db_connection.is_connected():
            self._db_connection = self.get_db_connection()
        
        if not self._db_connection:
            self.estadisticas['errores_bd'] += 1
            return {'status_level1': 'ERROR_BD', 'status_level2': 'ERROR_BD', 'found': False}
        
        try:
            cursor = self._db_connection.cursor(dictionary=True)
            
            # Buscar el teléfono en la base de datos
            # Usar la misma lógica que en procesar_llamadas_segurcaixa.py
            query = """
                SELECT telefono, status_level_1, status_level_2, updated_at
                FROM leads 
                WHERE REGEXP_REPLACE(telefono, '[^0-9]', '') = %s
                LIMIT 1
            """
            
            cursor.execute(query, (phone,))
            result = cursor.fetchone()
            cursor.close()
            
            if result:
                return {
                    'status_level1': result.get('status_level_1', ''),
                    'status_level2': result.get('status_level_2', ''),
                    'found': True,
                    'updated_at': result.get('updated_at', '')
                }
            else:
                return {'status_level1': 'NO_ENCONTRADO', 'status_level2': 'NO_ENCONTRADO', 'found': False}
                
        except mysql.connector.Error as err:
            logger.error(f"Error al consultar BD para teléfono {phone}: {err}")
            self.estadisticas['errores_bd'] += 1
            # Reintentar conexión en caso de error
            try:
                self._db_connection = self.get_db_connection()
            except:
                pass
            return {'status_level1': 'ERROR_BD', 'status_level2': 'ERROR_BD', 'found': False}
    
    def process_excel_file(self, file_path):
        """
        Procesar archivo Excel agregando columnas status_level1 y status_level2
        
        Args:
            file_path (str): Ruta al archivo Excel con URLs de grabación
        """
        logger.info(f"Iniciando procesamiento de: {file_path}")
        
        if not os.path.exists(file_path):
            logger.error(f"Archivo no encontrado: {file_path}")
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Encontrar la columna "To" que contiene el teléfono
            to_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and str(cell_value).lower() == 'to':
                    to_col = col
                    logger.info(f"Columna To (teléfono) encontrada: {col}")
                    break
            
            if not to_col:
                logger.error("No se encontró la columna 'To' con los teléfonos")
                return
            
            # Agregar columnas status_level_1 y status_level_2 al final
            status1_col = ws.max_column + 1
            status2_col = ws.max_column + 2
            
            ws.cell(row=1, column=status1_col, value="status_level_1")
            ws.cell(row=1, column=status2_col, value="status_level_2")
            
            logger.info(f"Agregadas columnas: status_level_1 (col {status1_col}), status_level_2 (col {status2_col})")
            
            # Procesar filas
            self.estadisticas['total_filas'] = ws.max_row - 1
            logger.info(f"Procesando {self.estadisticas['total_filas']} filas...")
            
            for row in range(2, ws.max_row + 1):
                # Obtener teléfono de la columna "To"
                phone_raw = ws.cell(row=row, column=to_col).value
                phone_normalized = self.normalize_phone(phone_raw)
                
                if phone_normalized:
                    logger.info(f"Fila {row}: Consultando teléfono {phone_normalized} (original: {phone_raw})")
                    status_data = self.get_status_from_db(phone_normalized)
                    
                    # Agregar los datos a las columnas
                    ws.cell(row=row, column=status1_col, value=status_data['status_level1'])
                    ws.cell(row=row, column=status2_col, value=status_data['status_level2'])
                    
                    if status_data['found']:
                        self.estadisticas['encontrados_bd'] += 1
                        logger.info(f"  [OK] Status: {status_data['status_level1']} -> {status_data['status_level2']}")
                    else:
                        self.estadisticas['no_encontrados'] += 1
                        logger.info(f"  [NO_ENCONTRADO] Teléfono no existe en BD")
                else:
                    # Sin teléfono válido
                    ws.cell(row=row, column=status1_col, value="SIN_TELEFONO")
                    ws.cell(row=row, column=status2_col, value="SIN_TELEFONO")
                    logger.info(f"Fila {row}: Sin teléfono válido (original: {phone_raw})")
                
                # Pausa para no saturar la BD
                if row % 50 == 0:
                    logger.info(f"  [PROGRESO] Procesadas {row-1} de {self.estadisticas['total_filas']} filas")
            
            # Guardar archivo con status
            output_file = file_path.replace(".xlsx", "_con_status.xlsx")
            wb.save(output_file)
            logger.info(f"Archivo guardado con status: {output_file}")
            
            self.print_statistics()
            
        except Exception as e:
            logger.error(f"Error al procesar archivo Excel: {e}")
        finally:
            # Cerrar la conexión compartida
            if hasattr(self, '_db_connection') and self._db_connection.is_connected():
                self._db_connection.close()
                logger.info("Conexión a Railway MySQL cerrada")
    
    def print_statistics(self):
        """Imprimir estadísticas finales"""
        logger.info("\n" + "="*60)
        logger.info("ESTADISTICAS FINALES - ACTUALIZACION DE STATUS")
        logger.info("="*60)
        logger.info(f"Total de filas procesadas: {self.estadisticas['total_filas']}")
        logger.info(f"Encontrados en BD: {self.estadisticas['encontrados_bd']}")
        logger.info(f"No encontrados en BD: {self.estadisticas['no_encontrados']}")
        logger.info(f"Errores de BD: {self.estadisticas['errores_bd']}")
        
        if self.estadisticas['total_filas'] > 0:
            tasa_encontrados = (self.estadisticas['encontrados_bd'] / self.estadisticas['total_filas']) * 100
            logger.info(f"Tasa de coincidencia: {tasa_encontrados:.2f}%")
        
        logger.info("="*60)


def main():
    """Función principal"""
    print("\n" + "="*60)
    print("AGREGADOR DE STATUS DESDE BASE DE DATOS DE RAILWAY")
    print("="*60)
    
    # Archivo por defecto (el que acabamos de generar)
    default_file = "Llamadas_07_30_2025_08_01_2025_procesar_grabaciones_campos_json.xlsx"
    
    print(f"\nUsando archivo por defecto: '{default_file}'")
    file_path = default_file
    
    # Verificar que existe
    if not os.path.exists(file_path):
        print(f"ERROR: Archivo no encontrado: {file_path}")
        return
    
    print(f"\nIniciando procesamiento...")
    print(f"Archivo: {file_path}")
    print("Conectando a base de datos de Railway...")
    
    updater = StatusUpdater()
    updater.process_excel_file(file_path)
    
    print(f"\nProcesamiento completado. Ver log: agregar_status_bd.log")


if __name__ == "__main__":
    main()