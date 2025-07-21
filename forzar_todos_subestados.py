#!/usr/bin/env python3
"""
Forzar TODOS los registros del Excel a tener subestados correctos
"""
import requests
import json
import time
from openpyxl import load_workbook

def extraer_todos_los_telefonos():
    """Extraer todos los teléfonos del Excel"""
    file_path = "C:\\Users\\jbeno\\CascadeProjects\\PearlInfo\\SegurcaixaCalls_julio18_procesar_grabaciones.xlsx"
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Encontrar columna CollectedInfo
        collected_info_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "collectedinfo" in str(cell_value).lower():
                collected_info_col = col
                break
        
        telefonos = []
        
        for row in range(2, ws.max_row + 1):
            collected_info_raw = ws.cell(row=row, column=collected_info_col).value
            
            try:
                if isinstance(collected_info_raw, str):
                    data = json.loads(collected_info_raw)
                elif isinstance(collected_info_raw, dict):
                    data = collected_info_raw
                else:
                    continue
                
                telefono = data.get('phoneNumber', '')
                nombre = f"{data.get('firstName', '')} {data.get('lastName', '')}".strip()
                
                # Limpiar teléfono
                if telefono.startswith('+34'):
                    telefono = telefono[3:]
                elif telefono.startswith('34') and len(telefono) == 11:
                    telefono = telefono[2:]
                
                if telefono:
                    # Determinar estado según flags JSON
                    if data.get('noInteresado', False):
                        estado = 'No Interesado'
                    else:
                        estado = 'Volver a llamar'  # Por defecto
                    
                    telefonos.append({
                        'telefono': telefono,
                        'nombre': nombre,
                        'estado_esperado': estado,
                        'fila': row
                    })
                    
            except Exception as e:
                print(f"Error procesando fila {row}: {e}")
                continue
        
        return telefonos
        
    except Exception as e:
        print(f"Error leyendo Excel: {e}")
        return []

def forzar_subestado(registro):
    """Forzar subestado correcto para un registro"""
    api_url = "https://tuotempo-apis-production.up.railway.app/api/actualizar_resultado"
    headers = {'Content-Type': 'application/json'}
    
    telefono = registro['telefono']
    estado = registro['estado_esperado']
    
    if estado == 'No Interesado':
        payload = {
            "telefono": telefono,
            "noInteresado": True,
            "razonNoInteres": "No da motivos - FORZADO",
            "codigoNoInteres": "otros"
        }
        esperado_level_2 = "No da motivos"
    else:  # Volver a llamar
        payload = {
            "telefono": telefono,
            "volverALlamar": True,
            "razonvueltaallamar": "no disponible cliente - FORZADO", 
            "codigoVolverLlamar": "interrupcion"
        }
        esperado_level_2 = "no disponible cliente"
    
    try:
        response = requests.post(api_url, json=payload, headers=headers, timeout=30)
        
        if response.status_code == 200:
            return True, esperado_level_2
        else:
            return False, f"Error {response.status_code}"
            
    except Exception as e:
        return False, f"Exception: {e}"

def main():
    """Función principal - forzar todos los subestados"""
    print("=== FORZADOR MASIVO DE SUBESTADOS ===")
    print("Este programa va a FORZAR todos los registros del Excel")
    print("a tener el subestado correcto, sin excepciones.")
    print()
    
    confirm = input("¿CONTINUAR? Esto actualizará todos los 39 registros (s/N): ").strip().lower()
    if confirm != 's':
        print("Cancelado.")
        return
    
    telefonos = extraer_todos_los_telefonos()
    
    if not telefonos:
        print("No se pudieron extraer teléfonos del Excel")
        return
    
    print(f"Procesando {len(telefonos)} registros...")
    print()
    
    exitosos = 0
    errores = 0
    no_interesado_count = 0
    volver_llamar_count = 0
    
    for i, registro in enumerate(telefonos, 1):
        print(f"{i:2}/39: {registro['nombre'][:25]:25} ({registro['telefono']}) -> {registro['estado_esperado']}")
        
        success, resultado = forzar_subestado(registro)
        
        if success:
            exitosos += 1
            if registro['estado_esperado'] == 'No Interesado':
                no_interesado_count += 1
            else:
                volver_llamar_count += 1
            print(f"      ✓ OK - Subestado: {resultado}")
        else:
            errores += 1
            print(f"      ✗ ERROR: {resultado}")
        
        time.sleep(0.3)  # Pausa para no saturar
    
    print(f"\n=== RESUMEN FINAL ===")
    print(f"Total procesados: {len(telefonos)}")
    print(f"Exitosos: {exitosos}")
    print(f"Errores: {errores}")
    print(f"No Interesado forzados: {no_interesado_count}")
    print(f"Volver a llamar forzados: {volver_llamar_count}")
    print()
    print("AHORA EL DASHBOARD DEBERÍA MOSTRAR:")
    print(f"- {volver_llamar_count} 'Volver a llamar' con {volver_llamar_count} subestados")
    print(f"- {no_interesado_count} 'No Interesado' con {no_interesado_count} subestados")
    print()
    print("Si aún faltan subestados, el problema está en:")
    print("1. El dashboard no muestra status_level_2 correctamente")
    print("2. Hay registros que no están en este Excel")
    print("3. La API tiene un bug que no detectamos")

if __name__ == "__main__":
    main()