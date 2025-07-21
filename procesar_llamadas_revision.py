"""
Procesador de Llamadas Segurcaixa - Modo Revisión Detallada

Este programa muestra el estado detectado para cada registro uno por uno
para permitir la validación manual antes de procesar masivamente.
"""

import os
import json
import logging
from openpyxl import load_workbook
from dotenv import load_dotenv
from mapeo_inteligente_segurcaixa import MapeadorInteligente

# Configurar logging simple
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

class RevisorSegurcaixa:
    def __init__(self):
        load_dotenv()
        self.mapeador = MapeadorInteligente()
        self.contador = 0
        self.por_estado = {}
    
    def extract_json_fields(self, collected_info_value):
        """Extraer campos del JSON collectedInfo"""
        fields = {
            # Campos de cliente
            "clinicaId": "", "certificado": "", "codigoPostal": "", "delegacion": "",
            "direccionClinica": "", "fechaNacimiento": "", "firstName": "", "lastName": "",
            "nif": "", "nombreClinica": "", "phoneNumber": "", "poliza": "", "segmento": "", "sexo": "",
            # Campos de resultado de llamada
            "noInteresado": False, "volverALlamar": False, "conPack": False, "sinPack": False,
            "fechaEscogida": "", "horaEscogida": "", "razonNoInteres": "", "razonVolverALlamar": ""
        }
        
        if not collected_info_value:
            return fields
        
        try:
            if isinstance(collected_info_value, str):
                data = json.loads(collected_info_value)
            elif isinstance(collected_info_value, dict):
                data = collected_info_value
            else:
                return fields
            
            for field in fields.keys():
                if field in data:
                    fields[field] = data[field]
        except:
            pass
        
        return fields
    
    def revisar_archivo(self, file_path):
        """Revisar archivo Excel registro por registro"""
        print(f"\n{'='*80}")
        print(f"REVISION DETALLADA: {os.path.basename(file_path)}")
        print(f"{'='*80}")
        
        if not os.path.exists(file_path):
            print(f"ERROR: Archivo no encontrado: {file_path}")
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Encontrar columnas
            collected_info_col = None
            call_id_col = None
            summary_col = None
            duration_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    cell_lower = str(cell_value).lower()
                    if "collectedinfo" in cell_lower:
                        collected_info_col = col
                    elif "id" in cell_lower and call_id_col is None:
                        call_id_col = col
                    elif "summary" in cell_lower or "resumen" in cell_lower:
                        summary_col = col
                    elif "duration" in cell_lower or "duracion" in cell_lower:
                        duration_col = col
            
            if not collected_info_col:
                print("ERROR: No se encontro la columna 'CollectedInfo'")
                return
            
            print(f"Columnas encontradas:")
            print(f"  - CollectedInfo: {collected_info_col}")
            print(f"  - Call ID: {call_id_col}")
            print(f"  - Summary: {summary_col}")
            print(f"  - Duration: {duration_col}")
            print()
            
            total_filas = ws.max_row - 1
            print(f"Total de registros a revisar: {total_filas}")
            print(f"{'='*80}")
            
            # Procesar cada fila
            for row in range(2, ws.max_row + 1):
                self.revisar_fila(ws, row, collected_info_col, call_id_col, summary_col, duration_col)
                
                # Cada 10 registros, preguntar si continuar
                if self.contador > 0 and self.contador % 10 == 0:
                    continuar = input(f"\n[Procesados {self.contador}/{total_filas}] Continuar? (s/N/q para salir): ").strip().lower()
                    if continuar == 'q':
                        print("Revision detenida por el usuario.")
                        break
                    elif continuar != 's':
                        print("Pausa en la revision.")
                        break
            
            self.mostrar_resumen()
            
        except Exception as e:
            print(f"ERROR al procesar archivo: {e}")
    
    def revisar_fila(self, worksheet, row_num, collected_info_col, call_id_col=None, summary_col=None, duration_col=None):
        """Revisar una fila individual"""
        try:
            self.contador += 1
            
            # Extraer datos
            collected_info_raw = worksheet.cell(row=row_num, column=collected_info_col).value
            call_id = worksheet.cell(row=row_num, column=call_id_col).value if call_id_col else ""
            summary = worksheet.cell(row=row_num, column=summary_col).value if summary_col else ""
            duration_raw = worksheet.cell(row=row_num, column=duration_col).value if duration_col else 0
            
            # Convertir duracion
            duracion = 0
            if duration_raw:
                try:
                    duracion = int(float(duration_raw))
                except:
                    duracion = 0
            
            print(f"\n{'-'*60}")
            print(f"REGISTRO #{self.contador} (Fila {row_num})")
            print(f"Call ID: {call_id}")
            print(f"Duracion: {duracion}s")
            print(f"{'-'*60}")
            
            # Extraer campos del JSON
            collected_info = self.extract_json_fields(collected_info_raw)
            
            # Mostrar datos del cliente
            print(f"CLIENTE:")
            print(f"  Nombre: {collected_info.get('firstName', 'N/A')} {collected_info.get('lastName', 'N/A')}")
            print(f"  Telefono: {collected_info.get('phoneNumber', 'N/A')}")
            print(f"  Clinica: {collected_info.get('nombreClinica', 'N/A')}")
            
            # Mostrar flags de resultado del JSON
            print(f"\nRESULTADO EN JSON:")
            print(f"  noInteresado: {collected_info.get('noInteresado', False)}")
            print(f"  conPack: {collected_info.get('conPack', False)}")
            print(f"  sinPack: {collected_info.get('sinPack', False)}")
            print(f"  volverALlamar: {collected_info.get('volverALlamar', False)}")
            if collected_info.get('fechaEscogida'):
                print(f"  fechaEscogida: {collected_info.get('fechaEscogida')}")
            if collected_info.get('horaEscogida'):
                print(f"  horaEscogida: {collected_info.get('horaEscogida')}")
            if collected_info.get('razonNoInteres'):
                print(f"  razonNoInteres: {collected_info.get('razonNoInteres')}")
            if collected_info.get('razonVolverALlamar'):
                print(f"  razonVolverALlamar: {collected_info.get('razonVolverALlamar')}")
            
            # Mostrar resumen si existe
            if summary:
                print(f"\nRESUMEN:")
                print(f"  {summary[:200]}{'...' if len(summary) > 200 else ''}")
            
            # Generar payload usando mapeo directo desde JSON
            payload = self.mapeador.mapear_a_api_payload_desde_json(collected_info)
            
            # Determinar estado basado en el payload generado
            if payload:
                if payload.get('noInteresado'):
                    nivel_1, nivel_2 = 'No Interesado', 'No da motivos'
                elif payload.get('conPack'):
                    nivel_1, nivel_2 = 'CONFIRMADO', 'Con Pack'
                elif payload.get('nuevaCita'):
                    nivel_1, nivel_2 = 'CONFIRMADO', 'Sin Pack'
                elif payload.get('volverALlamar'):
                    nivel_1, nivel_2 = 'Volver a llamar', 'no disponible cliente'
                else:
                    nivel_1, nivel_2 = 'Sin clasificar', 'Sin clasificar'
            else:
                nivel_1, nivel_2 = 'Error', 'Error'
            
            print(f"\nESTADO FINAL:")
            print(f"  Nivel 1: {nivel_1}")
            print(f"  Nivel 2: {nivel_2}")
            
            if payload:
                print(f"\nPAYLOAD PARA API:")
                for key, value in payload.items():
                    if key != 'telefono':  # No mostrar telefono por privacidad
                        print(f"  {key}: {value}")
                
                # Validar payload
                es_valido, errores = self.mapeador.validar_payload(payload)
                if es_valido:
                    print(f"  ✓ Payload VALIDO")
                    
                    # Contar por estado
                    estado_key = f"{nivel_1} -> {nivel_2}"
                    if estado_key not in self.por_estado:
                        self.por_estado[estado_key] = 0
                    self.por_estado[estado_key] += 1
                else:
                    print(f"  ✗ Payload INVALIDO: {errores}")
            else:
                print(f"  ✗ No se pudo generar payload")
            
        except Exception as e:
            print(f"ERROR procesando fila {row_num}: {e}")
    
    def mostrar_resumen(self):
        """Mostrar resumen final"""
        print(f"\n{'='*80}")
        print(f"RESUMEN DE LA REVISION")
        print(f"{'='*80}")
        print(f"Total registros procesados: {self.contador}")
        
        if self.por_estado:
            print(f"\nDistribucion por estado:")
            for estado, cantidad in sorted(self.por_estado.items()):
                porcentaje = (cantidad / self.contador) * 100 if self.contador > 0 else 0
                print(f"  {estado}: {cantidad} ({porcentaje:.1f}%)")
        
        print(f"{'='*80}")

def main():
    """Funcion principal"""
    import sys
    
    print("REVISOR DE ESTADOS - LLAMADAS SEGURCAIXA")
    print("Este programa muestra el estado detectado para cada registro")
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = input("\nRuta del archivo Excel: ").strip()
        if not file_path:
            file_path = "C:\\Users\\jbeno\\CascadeProjects\\PearlInfo\\SegurcaixaCalls_julio18_procesar_grabaciones.xlsx"
    
    revisor = RevisorSegurcaixa()
    revisor.revisar_archivo(file_path)

if __name__ == "__main__":
    main()