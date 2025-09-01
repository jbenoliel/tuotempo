#!/usr/bin/env python3
"""
Contar cuántas URLs de grabación reales hay en el archivo Excel
"""

from openpyxl import load_workbook
import os

def contar_urls():
    file_path = r"C:\Users\jbeno\CascadeProjects\tuotempo\Llamadas_07_30_2025_08_01_2025_procesar_grabaciones_campos_json.xlsx"
    
    print(f"Contando URLs en: {os.path.basename(file_path)}")
    print("="*60)
    
    if not os.path.exists(file_path):
        print(f"ERROR: Archivo no encontrado: {file_path}")
        return
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        total_filas = ws.max_row - 1  # Excluyendo header
        urls_disponibles = 0
        urls_no_disponibles = 0
        urls_vacias = 0
        
        print("Analizando URLs de grabación...")
        
        for row in range(2, ws.max_row + 1):  # Empezar desde fila 2 (datos)
            url_value = ws.cell(row=row, column=2).value  # Columna 2 es "Recording URL"
            
            if url_value is None or url_value == "":
                urls_vacias += 1
            elif str(url_value).lower() == "no disponible":
                urls_no_disponibles += 1
            else:
                urls_disponibles += 1
                if urls_disponibles <= 5:  # Mostrar las primeras 5 URLs encontradas
                    call_id = ws.cell(row=row, column=1).value
                    print(f"URL encontrada en fila {row} (Call ID: {call_id}): {url_value}")
        
        print("\n" + "="*60)
        print("RESUMEN:")
        print(f"Total de filas de datos: {total_filas}")
        print(f"URLs disponibles: {urls_disponibles}")
        print(f"URLs 'No disponible': {urls_no_disponibles}")
        print(f"URLs vacías: {urls_vacias}")
        print(f"Porcentaje con URL: {(urls_disponibles/total_filas)*100:.2f}%")
        
    except Exception as e:
        print(f"ERROR: {e}")

if __name__ == "__main__":
    contar_urls()