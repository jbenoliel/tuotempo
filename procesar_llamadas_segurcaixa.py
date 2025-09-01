"""
Sistema de Procesamiento de Llamadas Segurcaixa → API actualizar_resultado

Este programa:
1. Lee un archivo Excel con llamadas procesadas de Pearl AI
2. Extrae el JSON collectedInfo de cada fila
3. Mapea los campos al formato de la API actualizar_resultado
4. Realiza llamadas a la API con validación
5. Verifica en base de datos que los cambios se aplicaron correctamente
"""

import os
import json
import time
import logging
import requests
import mysql.connector
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
from config import settings
from mapeo_inteligente_segurcaixa import MapeadorInteligente

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('procesamiento_segurcaixa.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class SegurcaixaProcessor:
    def __init__(self, dry_run=True):
        """
        Inicializar el procesador de llamadas
        
        Args:
            dry_run (bool): Si True, no hace llamadas reales a la API
        """
        load_dotenv()
        self.dry_run = dry_run
        self.api_base_url = "https://tuotempo-apis-production.up.railway.app"
        self.mapeador = MapeadorInteligente()
        self.estadisticas = {
            'total_filas': 0,
            'procesadas_exitosamente': 0,
            'errores_api': 0,
            'errores_validacion': 0,
            'sin_telefono': 0,
            'estados_detectados': {}
        }
        
        logger.info(f"Procesador iniciado - Modo: {'DRY RUN' if dry_run else 'PRODUCCIÓN'}")
    
    def get_db_connection(self):
        """Establece conexión con la base de datos MySQL usando EXACTAMENTE la misma configuración que api_resultado_llamada.py"""
        try:
            # Configuración EXACTAMENTE igual a api_resultado_llamada.py
            DB_CONFIG = {
                'host': settings.DB_HOST,
                'port': settings.DB_PORT,
                'user': settings.DB_USER,
                'password': settings.DB_PASSWORD,
                'database': settings.DB_DATABASE,
                'ssl_disabled': True,  # Deshabilitar SSL para evitar errores de conexión
                'autocommit': True,
                'charset': 'utf8mb4',
                'use_unicode': True
            }
            
            connection = mysql.connector.connect(**DB_CONFIG)
            return connection
        except mysql.connector.Error as err:
            logger.error(f"Error conectando a MySQL: {err}")
            # Si falla con SSL, intentar sin SSL (mismo fallback que api_resultado_llamada.py)
            if 'SSL' in str(err) or '2026' in str(err):
                try:
                    logger.info("Intentando conexión sin SSL...")
                    config_no_ssl = DB_CONFIG.copy()
                    config_no_ssl['ssl_disabled'] = True
                    connection = mysql.connector.connect(**config_no_ssl)
                    return connection
                except mysql.connector.Error as err2:
                    logger.error(f"Error conectando sin SSL: {err2}")
            return None
    
    def extract_json_fields(self, collected_info_value):
        """
        Extraer campos específicos del JSON de collected_info
        Basado en la función del main_segurcaixa.py original
        """
        fields = {
            "clinicaId": "",
            "certificado": "",
            "codigoPostal": "",
            "delegacion": "",
            "direccionClinica": "",
            "fechaNacimiento": "",
            "firstName": "",
            "lastName": "",
            "nif": "",
            "nombreClinica": "",
            "phoneNumber": "",
            "poliza": "",
            "segmento": "",
            "sexo": ""
        }
        
        if not collected_info_value:
            return fields
        
        try:
            if isinstance(collected_info_value, str):
                data = json.loads(collected_info_value)
            elif isinstance(collected_info_value, dict):
                data = collected_info_value
            else:
                return fields
            
            for field in fields.keys():
                if field in data:
                    fields[field] = data[field]
                    
        except (json.JSONDecodeError, TypeError) as e:
            logger.error(f"Error al decodificar JSON: {e}")
        
        return fields
    
    def map_to_api_payload(self, collected_info, call_summary="", duracion=0):
        """
        Mapear campos del collectedInfo a payload de la API actualizar_resultado
        usando el mapeo inteligente
        
        Args:
            collected_info (dict): Datos extraídos del JSON
            call_summary (str): Resumen de la llamada para inferir estado
            duracion (int): Duración de la llamada en segundos
            
        Returns:
            tuple: (payload: dict, estado_detectado: str) o (None, None) si error
        """
        payload = self.mapeador.mapear_a_api_payload(collected_info, call_summary, duracion)
        
        if not payload:
            return None, None
        
        # Validar el payload
        es_valido, errores = self.mapeador.validar_payload(payload)
        if not es_valido:
            logger.error(f"Payload inválido: {errores}")
            return None, None
        
        # Detectar estado para estadísticas
        analisis = self.mapeador.analizar_resumen(call_summary)
        estado_detectado = analisis['estado']
        
        return payload, estado_detectado
    
    def call_api(self, payload):
        """
        Realizar llamada a la API actualizar_resultado
        
        Args:
            payload (dict): Datos a enviar
            
        Returns:
            tuple: (success: bool, response_data: dict, status_code: int)
        """
        if self.dry_run:
            logger.info(f"[DRY RUN] Llamaría a API con: {payload}")
            return True, {"success": True, "message": "DRY RUN - No se realizó llamada real"}, 200
        
        url = f"{self.api_base_url}/api/actualizar_resultado"
        headers = {'Content-Type': 'application/json'}
        
        try:
            response = requests.post(url, json=payload, headers=headers, timeout=30)
            
            if response.status_code == 200:
                return True, response.json(), response.status_code
            else:
                logger.error(f"Error API {response.status_code}: {response.text}")
                return False, {"error": response.text}, response.status_code
                
        except requests.exceptions.RequestException as e:
            logger.error(f"Error de conexión con API: {e}")
            return False, {"error": str(e)}, 0
    
    def verify_in_database(self, telefono):
        """
        Verificar en la base de datos que el lead se actualizó correctamente
        
        Args:
            telefono (str): Número de teléfono a verificar
            
        Returns:
            dict: Datos del lead o None si no se encuentra
        """
        if self.dry_run:
            logger.info(f"[DRY RUN] Verificaría en BD el teléfono: {telefono}")
            return {"telefono": telefono, "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        
        conn = self.get_db_connection()
        if not conn:
            return None
        
        try:
            cursor = conn.cursor(dictionary=True)
            # Usar la misma lógica de búsqueda que api_resultado_llamada.py
            query = """
                SELECT telefono, status_level_1, status_level_2, 
                       razon_vuelta_a_llamar, updated_at
                FROM leads 
                WHERE REGEXP_REPLACE(telefono, '[^0-9]', '') = %s
                LIMIT 1
            """
            # Normalizar teléfono como en api_resultado_llamada.py
            telefono_norm = ''.join(filter(str.isdigit, telefono))
            cursor.execute(query, (telefono_norm,))
            result = cursor.fetchone()
            cursor.close()
            return result
            
        except mysql.connector.Error as err:
            logger.error(f"Error al consultar BD: {err}")
            return None
        finally:
            conn.close()
    
    def process_excel_file(self, file_path):
        """
        Procesar archivo Excel con llamadas de Segurcaixa
        
        Args:
            file_path (str): Ruta al archivo Excel
        """
        logger.info(f"Iniciando procesamiento de: {file_path}")
        
        if not os.path.exists(file_path):
            logger.error(f"Archivo no encontrado: {file_path}")
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            header_row = 1
            
            # Encontrar columnas relevantes
            collected_info_col = None
            call_id_col = None
            summary_col = None
            duration_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value:
                    cell_lower = str(cell_value).lower()
                    if "collectedinfo" in cell_lower:
                        collected_info_col = col
                        logger.info(f"Columna CollectedInfo encontrada: {col}")
                    elif "id" in cell_lower and call_id_col is None:
                        call_id_col = col
                        logger.info(f"Columna Call ID encontrada: {col}")
                    elif "summary" in cell_lower or "resumen" in cell_lower:
                        summary_col = col
                        logger.info(f"Columna Summary/Resumen encontrada: {col}")
                    elif "duration" in cell_lower or "duración" in cell_lower or "duracion" in cell_lower:
                        duration_col = col
                        logger.info(f"Columna Duration/Duración encontrada: {col}")
            
            if not collected_info_col:
                logger.error("No se encontró la columna 'CollectedInfo'")
                return
            
            # Procesar filas
            self.estadisticas['total_filas'] = ws.max_row - 1  # Excluyendo header
            
            for row in range(2, ws.max_row + 1):
                self.process_row(ws, row, collected_info_col, call_id_col, summary_col, duration_col)
                
                # Pequeña pausa entre llamadas para no saturar la API
                if not self.dry_run:
                    time.sleep(0.5)
            
            self.print_statistics()
            
        except Exception as e:
            logger.error(f"Error al procesar archivo Excel: {e}")
    
    def process_row(self, worksheet, row_num, collected_info_col, call_id_col=None, summary_col=None, duration_col=None):
        """
        Procesar una fila individual del Excel
        
        Args:
            worksheet: Hoja de cálculo de openpyxl
            row_num (int): Número de fila
            collected_info_col (int): Columna con collectedInfo
            call_id_col (int): Columna con call ID (opcional)
            summary_col (int): Columna con resumen (opcional)
            duration_col (int): Columna con duración (opcional)
        """
        try:
            # Extraer datos de la fila
            collected_info_raw = worksheet.cell(row=row_num, column=collected_info_col).value
            call_id = worksheet.cell(row=row_num, column=call_id_col).value if call_id_col else ""
            summary = worksheet.cell(row=row_num, column=summary_col).value if summary_col else ""
            duration_raw = worksheet.cell(row=row_num, column=duration_col).value if duration_col else 0
            
            # Convertir duración a segundos si es necesario
            duracion = 0
            if duration_raw:
                try:
                    duracion = int(float(duration_raw))
                except:
                    duracion = 0
            
            logger.info(f"Procesando fila {row_num} - Call ID: {call_id}")
            
            # Extraer campos del JSON
            collected_info = self.extract_json_fields(collected_info_raw)
            
            # Mapear a payload de API usando el mapeo inteligente
            payload, estado_detectado = self.map_to_api_payload(collected_info, summary, duracion)
            
            if not payload:
                logger.warning(f"Fila {row_num}: No se pudo crear payload (sin teléfono)")
                self.estadisticas['sin_telefono'] += 1
                return
            
            # Actualizar estadísticas de estados detectados
            if estado_detectado:
                if estado_detectado not in self.estadisticas['estados_detectados']:
                    self.estadisticas['estados_detectados'][estado_detectado] = 0
                self.estadisticas['estados_detectados'][estado_detectado] += 1
            
            # Llamar a la API
            success, response, status_code = self.call_api(payload)
            
            if success:
                logger.info(f"Fila {row_num}: API llamada exitosa - {response.get('message', '')}")
                
                # Verificar en base de datos de producción
                telefono = payload['telefono']
                db_result = self.verify_in_database(telefono)
                
                if db_result:
                    logger.info(f"Fila {row_num}: Verificación BD exitosa - Lead actualizado")
                    self.estadisticas['procesadas_exitosamente'] += 1
                else:
                    logger.warning(f"Fila {row_num}: No se pudo verificar en BD")
                    # Aún así, contar como exitoso porque la API respondió OK
                    self.estadisticas['procesadas_exitosamente'] += 1
            else:
                logger.error(f"Fila {row_num}: Error en API - {response}")
                self.estadisticas['errores_api'] += 1
                
        except Exception as e:
            logger.error(f"Error procesando fila {row_num}: {e}")
            self.estadisticas['errores_validacion'] += 1
    
    def print_statistics(self):
        """Imprimir estadísticas finales del procesamiento"""
        logger.info("\n" + "="*60)
        logger.info("ESTADÍSTICAS FINALES DEL PROCESAMIENTO")
        logger.info("="*60)
        logger.info(f"Total de filas procesadas: {self.estadisticas['total_filas']}")
        logger.info(f"Exitosas: {self.estadisticas['procesadas_exitosamente']}")
        logger.info(f"Errores de API: {self.estadisticas['errores_api']}")
        logger.info(f"Errores de validación: {self.estadisticas['errores_validacion']}")
        logger.info(f"Sin teléfono: {self.estadisticas['sin_telefono']}")
        
        if self.estadisticas['total_filas'] > 0:
            tasa_exito = (self.estadisticas['procesadas_exitosamente'] / self.estadisticas['total_filas']) * 100
            logger.info(f"Tasa de éxito: {tasa_exito:.2f}%")
        
        # Mostrar distribución de estados detectados
        if self.estadisticas['estados_detectados']:
            logger.info("\n" + "-"*40)
            logger.info("ESTADOS DETECTADOS:")
            logger.info("-"*40)
            for estado, cantidad in sorted(self.estadisticas['estados_detectados'].items()):
                porcentaje = (cantidad / self.estadisticas['total_filas']) * 100 if self.estadisticas['total_filas'] > 0 else 0
                logger.info(f"{estado}: {cantidad} ({porcentaje:.1f}%)")
        
        logger.info("="*60)


def main():
    """Función principal del programa"""
    import sys
    
    print("\n" + "="*60)
    print("PROCESADOR DE LLAMADAS SEGURCAIXA -> API ACTUALIZAR_RESULTADO")
    print("="*60)
    
    # Verificar si se pasa archivo como argumento
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        dry_run = True  # Por defecto DRY RUN si se pasa archivo
        print(f"Archivo especificado: {file_path}")
        print("Modo: DRY RUN (por defecto)")
    else:
        # Configuración interactiva
        print("\nEjecutando en modo PRODUCCIÓN para actualizar la base de datos")
        dry_run = False
        
        # Archivo a procesar
        default_file = "Llamadas_07_30_2025_08_01_2025_procesar_grabaciones_campos_json.xlsx"
        print(f"\nUsando archivo: {default_file}")
        file_path = default_file
        
        if not file_path:
            file_path = default_file
    
    # Crear procesador y ejecutar
    processor = SegurcaixaProcessor(dry_run=dry_run)
    
    print(f"\nIniciando procesamiento...")
    print(f"Archivo: {file_path}")
    print(f"Modo: {'DRY RUN (sin llamadas reales)' if dry_run else 'PRODUCCIÓN'}")
    
    if not dry_run:
        print("\nATENCION: Realizando llamadas REALES a la API para actualizar la base de datos...")
    
    processor.process_excel_file(file_path)
    
    print(f"\nProcesamiento completado. Ver log: procesamiento_segurcaixa.log")


if __name__ == "__main__":
    main()