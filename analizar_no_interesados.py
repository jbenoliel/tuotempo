#!/usr/bin/env python3
"""
Análisis detallado de todos los registros No Interesado
"""
import os
import json
import logging
from openpyxl import load_workbook
from dotenv import load_dotenv

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

class AnalizadorNoInteresados:
    def __init__(self):
        load_dotenv()
        
    def extract_json_fields(self, collected_info_value):
        """Extraer campos del JSON collectedInfo"""
        fields = {
            "firstName": "", "lastName": "", "phoneNumber": "", "nombreClinica": "",
            "noInteresado": False, "volverALlamar": False, "conPack": False, "sinPack": False,
            "razonNoInteres": ""
        }
        
        if not collected_info_value:
            return fields
        
        try:
            if isinstance(collected_info_value, str):
                data = json.loads(collected_info_value)
            elif isinstance(collected_info_value, dict):
                data = collected_info_value
            else:
                return fields
            
            for field in fields.keys():
                if field in data:
                    fields[field] = data[field]
        except:
            pass
        
        return fields
    
    def analizar_archivo(self, file_path):
        """Analizar todos los registros del archivo"""
        print("=== ANÁLISIS COMPLETO DE REGISTROS ===")
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Encontrar columna CollectedInfo
        collected_info_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "collectedinfo" in str(cell_value).lower():
                collected_info_col = col
                break
        
        if not collected_info_col:
            print("ERROR: No se encontró columna CollectedInfo")
            return
        
        # Contadores
        contadores = {
            'no_interesado': 0,
            'volver_a_llamar': 0,
            'con_pack': 0,
            'sin_pack': 0,
            'sin_clasificar': 0
        }
        
        # Listas para análisis detallado
        registros_no_interesado = []
        registros_sin_clasificar = []
        
        print(f"Analizando {ws.max_row - 1} registros...\n")
        
        for row in range(2, ws.max_row + 1):
            collected_info_raw = ws.cell(row=row, column=collected_info_col).value
            collected_info = self.extract_json_fields(collected_info_raw)
            
            nombre = f"{collected_info.get('firstName', '')} {collected_info.get('lastName', '')}".strip()
            telefono = collected_info.get('phoneNumber', '')
            
            # Limpiar teléfono
            tel_limpio = telefono
            if telefono.startswith('+34'):
                tel_limpio = telefono[3:]
            elif telefono.startswith('34') and len(telefono) == 11:
                tel_limpio = telefono[2:]
            
            # Clasificar según flags JSON
            if collected_info.get('noInteresado', False):
                contadores['no_interesado'] += 1
                registros_no_interesado.append({
                    'fila': row,
                    'nombre': nombre,
                    'telefono': tel_limpio,
                    'telefono_original': telefono,
                    'razon': collected_info.get('razonNoInteres', 'Sin razón')
                })
            elif collected_info.get('conPack', False):
                contadores['con_pack'] += 1
            elif collected_info.get('sinPack', False):
                contadores['sin_pack'] += 1
            elif collected_info.get('volverALlamar', False):
                contadores['volver_a_llamar'] += 1
            else:
                contadores['sin_clasificar'] += 1
                registros_sin_clasificar.append({
                    'fila': row,
                    'nombre': nombre,
                    'telefono': tel_limpio,
                    'flags': {
                        'noInteresado': collected_info.get('noInteresado', False),
                        'conPack': collected_info.get('conPack', False),
                        'sinPack': collected_info.get('sinPack', False),
                        'volverALlamar': collected_info.get('volverALlamar', False)
                    }
                })
        
        # Mostrar resumen
        print("RESUMEN POR CLASIFICACIÓN:")
        print(f"  No Interesado: {contadores['no_interesado']}")
        print(f"  Volver a llamar: {contadores['volver_a_llamar']}")  
        print(f"  Con Pack: {contadores['con_pack']}")
        print(f"  Sin Pack: {contadores['sin_pack']}")
        print(f"  Sin clasificar: {contadores['sin_clasificar']}")
        print()
        
        # Detalles de No Interesado
        print("DETALLE DE REGISTROS 'NO INTERESADO':")
        print("-" * 80)
        for i, reg in enumerate(registros_no_interesado, 1):
            print(f"{i:2}. Fila {reg['fila']:2} | {reg['nombre']:30} | {reg['telefono']:9} | {reg['razon']}")
        
        print(f"\nTotal registros No Interesado encontrados: {len(registros_no_interesado)}")
        
        if registros_sin_clasificar:
            print("\nREGISTROS SIN CLASIFICAR (pueden ser No Interesado adicionales):")
            print("-" * 80)
            for i, reg in enumerate(registros_sin_clasificar, 1):
                print(f"{i:2}. Fila {reg['fila']:2} | {reg['nombre']:30} | {reg['telefono']:9}")
                print(f"    Flags: {reg['flags']}")
        
        print("\n" + "=" * 80)
        
        # Verificar si hay registros que deberían ser No Interesado pero no tienen el flag
        posibles_no_interesado = []
        for row in range(2, ws.max_row + 1):
            collected_info_raw = ws.cell(row=row, column=collected_info_col).value
            collected_info = self.extract_json_fields(collected_info_raw)
            
            # Si no tiene ningún flag True pero tiene razón de no interés
            if (not any([
                collected_info.get('noInteresado', False),
                collected_info.get('conPack', False), 
                collected_info.get('sinPack', False),
                collected_info.get('volverALlamar', False)
            ]) and collected_info.get('razonNoInteres', '')):
                
                nombre = f"{collected_info.get('firstName', '')} {collected_info.get('lastName', '')}".strip()
                telefono = collected_info.get('phoneNumber', '')
                posibles_no_interesado.append({
                    'fila': row,
                    'nombre': nombre,
                    'telefono': telefono,
                    'razon': collected_info.get('razonNoInteres', '')
                })
        
        if posibles_no_interesado:
            print("POSIBLES NO INTERESADO ADICIONALES (con razón pero sin flag):")
            print("-" * 80)
            for reg in posibles_no_interesado:
                print(f"Fila {reg['fila']:2} | {reg['nombre']:30} | {reg['telefono']:15} | {reg['razon']}")

def main():
    file_path = "C:\\Users\\jbeno\\CascadeProjects\\PearlInfo\\SegurcaixaCalls_julio18_procesar_grabaciones.xlsx"
    
    analizador = AnalizadorNoInteresados()
    analizador.analizar_archivo(file_path)

if __name__ == "__main__":
    main()