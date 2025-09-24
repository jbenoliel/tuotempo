"""
Script para actualizar status_level_1 y status_level_2 desde el archivo Excel de Septiembre
"""

import pandas as pd
import pymysql
from datetime import datetime
from dotenv import load_dotenv
import os

load_dotenv()

def get_db_connection():
    """Crear conexión a la base de datos Railway"""
    try:
        connection = pymysql.connect(
            host=os.getenv('MYSQLHOST'),
            port=int(os.getenv('MYSQLPORT')),
            user=os.getenv('MYSQLUSER'),
            password=os.getenv('MYSQLPASSWORD'),
            database=os.getenv('MYSQLDATABASE'),
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        return connection
    except Exception as e:
        print(f"Error conectando a la base de datos: {e}")
        return None

def actualizar_septiembre():
    """Actualizar status de Septiembre"""

    # Usar el archivo v2 con los cambios
    archivo_excel = 'leads_Septiembre_20250922_063923v2.xlsx'

    print("ACTUALIZANDO STATUS DE SEPTIEMBRE")
    print("=" * 35)
    print(f"Archivo: {archivo_excel}")

    try:
        # Intentar diferentes métodos de lectura
        df = None

        # Método 1: Sin filtros
        try:
            df = pd.read_excel(archivo_excel, engine='openpyxl')
        except:
            pass

        # Método 2: Leer solo datos (ignorar filtros)
        if df is None:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(archivo_excel, data_only=True)
                ws = wb.active
                data = []
                headers = None

                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = row
                    else:
                        data.append(row)

                df = pd.DataFrame(data, columns=headers)
                wb.close()
            except Exception as e2:
                print(f"  Error método 2: {e2}")

        # Método 3: Intentar sin engine específico
        if df is None:
            try:
                df = pd.read_excel(archivo_excel, sheet_name=0)
            except Exception as e3:
                print(f"  Error método 3: {e3}")

        # Método 4: Cargar sin filtros automáticos
        if df is None:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(archivo_excel, read_only=True, data_only=True)
                ws = wb.active

                # Leer todas las filas
                all_rows = list(ws.iter_rows(values_only=True))
                if all_rows:
                    headers = all_rows[0]
                    data = all_rows[1:]
                    df = pd.DataFrame(data, columns=headers)

                wb.close()
                print("  Éxito con método 4 (read_only)")
            except Exception as e4:
                print(f"  Error método 4: {e4}")

        if df is None:
            print("ERROR: No se pudo leer el archivo con ningún método")
            return

        print(f"Registros leidos: {len(df)}")
        print(f"Primeras 5 columnas: {list(df.columns)[:5]}")

        # Verificar columnas
        if 'id' not in df.columns or 'status_level_1' not in df.columns:
            print("ERROR: Faltan columnas necesarias")
            return

        # Conectar a BD
        connection = get_db_connection()
        if not connection:
            return

        with connection.cursor() as cursor:
            cambios_aplicados = 0
            errores = 0

            print("Procesando cambios...")

            for _, row in df.iterrows():
                try:
                    lead_id = int(row['id'])
                    nuevo_status_1 = str(row['status_level_1']) if pd.notna(row['status_level_1']) else None
                    nuevo_status_2 = str(row['status_level_2']) if pd.notna(row['status_level_2']) else None

                    # Obtener datos actuales
                    cursor.execute("""
                        SELECT status_level_1, status_level_2, nombre
                        FROM leads
                        WHERE id = %s AND origen_archivo = 'Septiembre'
                    """, (lead_id,))

                    lead_actual = cursor.fetchone()

                    if lead_actual:
                        actual_1 = str(lead_actual['status_level_1']) if lead_actual['status_level_1'] else None
                        actual_2 = str(lead_actual['status_level_2']) if lead_actual['status_level_2'] else None

                        # Verificar cambios
                        if nuevo_status_1 != actual_1 or nuevo_status_2 != actual_2:
                            # Actualizar
                            cursor.execute("""
                                UPDATE leads
                                SET status_level_1 = %s, status_level_2 = %s, updated_at = %s
                                WHERE id = %s
                            """, (nuevo_status_1, nuevo_status_2, datetime.now(), lead_id))

                            cambios_aplicados += 1

                            if cambios_aplicados <= 5:  # Mostrar primeros 5 cambios
                                print(f"  ID {lead_id} ({lead_actual['nombre']}):")
                                if nuevo_status_1 != actual_1:
                                    print(f"    status_level_1: '{actual_1}' -> '{nuevo_status_1}'")
                                if nuevo_status_2 != actual_2:
                                    print(f"    status_level_2: '{actual_2}' -> '{nuevo_status_2}'")

                            if cambios_aplicados % 50 == 0:
                                print(f"  Procesados {cambios_aplicados} cambios...")

                except Exception as e:
                    errores += 1
                    if errores <= 3:  # Mostrar primeros errores
                        print(f"  Error en fila {row.name}: {e}")

            # Confirmar cambios
            connection.commit()

            print(f"\nRESULTADO:")
            print(f"  Cambios aplicados: {cambios_aplicados}")
            print(f"  Errores: {errores}")

            # Verificar estadísticas finales
            cursor.execute("""
                SELECT
                    status_level_1,
                    COUNT(*) as cantidad
                FROM leads
                WHERE origen_archivo = 'Septiembre'
                GROUP BY status_level_1
                ORDER BY cantidad DESC
                LIMIT 10
            """)

            print(f"\nEstadisticas de status_level_1 actualizadas:")
            for stat in cursor.fetchall():
                print(f"  {stat['status_level_1']}: {stat['cantidad']} leads")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if 'connection' in locals():
            connection.close()

if __name__ == "__main__":
    actualizar_septiembre()