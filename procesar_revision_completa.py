"""
Procesador de Llamadas Segurcaixa - Revisión Completa

Procesa todos los registros sin pausas para generar el reporte completo.
"""

import os
import json
import logging
from openpyxl import load_workbook
from dotenv import load_dotenv
from mapeo_inteligente_segurcaixa import MapeadorInteligente

# Configurar logging simple
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

class RevisorCompleto:
    def __init__(self):
        load_dotenv()
        self.mapeador = MapeadorInteligente()
        self.contador = 0
        self.por_estado = {}
        self.registros_detallados = []
    
    def extract_json_fields(self, collected_info_value):
        """Extraer campos del JSON collectedInfo"""
        fields = {
            # Campos de cliente
            "clinicaId": "", "certificado": "", "codigoPostal": "", "delegacion": "",
            "direccionClinica": "", "fechaNacimiento": "", "firstName": "", "lastName": "",
            "nif": "", "nombreClinica": "", "phoneNumber": "", "poliza": "", "segmento": "", "sexo": "",
            # Campos de resultado de llamada
            "noInteresado": False, "volverALlamar": False, "conPack": False, "sinPack": False,
            "fechaEscogida": "", "horaEscogida": "", "razonNoInteres": "", "razonVolverALlamar": ""
        }
        
        if not collected_info_value:
            return fields
        
        try:
            if isinstance(collected_info_value, str):
                data = json.loads(collected_info_value)
            elif isinstance(collected_info_value, dict):
                data = collected_info_value
            else:
                return fields
            
            for field in fields.keys():
                if field in data:
                    fields[field] = data[field]
        except:
            pass
        
        return fields
    
    def procesar_archivo_completo(self, file_path):
        """Procesar archivo Excel completo"""
        print(f"REVISION COMPLETA: {os.path.basename(file_path)}")
        print("=" * 80)
        
        if not os.path.exists(file_path):
            print(f"ERROR: Archivo no encontrado: {file_path}")
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Encontrar columnas
            collected_info_col = None
            call_id_col = None
            summary_col = None
            duration_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    cell_lower = str(cell_value).lower()
                    if "collectedinfo" in cell_lower:
                        collected_info_col = col
                    elif "id" in cell_lower and call_id_col is None:
                        call_id_col = col
                    elif "summary" in cell_lower or "resumen" in cell_lower:
                        summary_col = col
                    elif "duration" in cell_lower or "duracion" in cell_lower:
                        duration_col = col
            
            if not collected_info_col:
                print("ERROR: No se encontro la columna 'CollectedInfo'")
                return
            
            total_filas = ws.max_row - 1
            print(f"Procesando {total_filas} registros...")
            print()
            
            # Procesar todas las filas
            for row in range(2, ws.max_row + 1):
                self.procesar_fila(ws, row, collected_info_col, call_id_col, summary_col, duration_col)
            
            self.mostrar_resumen_final()
            
        except Exception as e:
            print(f"ERROR al procesar archivo: {e}")
    
    def procesar_fila(self, worksheet, row_num, collected_info_col, call_id_col=None, summary_col=None, duration_col=None):
        """Procesar una fila individual"""
        try:
            self.contador += 1
            
            # Extraer datos
            collected_info_raw = worksheet.cell(row=row_num, column=collected_info_col).value
            call_id = worksheet.cell(row=row_num, column=call_id_col).value if call_id_col else ""
            summary = worksheet.cell(row=row_num, column=summary_col).value if summary_col else ""
            
            # Extraer campos del JSON
            collected_info = self.extract_json_fields(collected_info_raw)
            
            # Generar payload usando mapeo directo desde JSON
            payload = self.mapeador.mapear_a_api_payload_desde_json(collected_info)
            
            # Determinar estado basado en el payload generado
            if payload:
                if payload.get('noInteresado'):
                    nivel_1, nivel_2 = 'No Interesado', 'No da motivos'
                elif payload.get('conPack'):
                    nivel_1, nivel_2 = 'CONFIRMADO', 'Con Pack'
                elif payload.get('nuevaCita'):
                    nivel_1, nivel_2 = 'CONFIRMADO', 'Sin Pack'
                elif payload.get('volverALlamar'):
                    nivel_1, nivel_2 = 'Volver a llamar', 'no disponible cliente'
                else:
                    nivel_1, nivel_2 = 'Sin clasificar', 'Sin clasificar'
            else:
                nivel_1, nivel_2 = 'Error', 'Error'
            
            # Guardar registro detallado
            registro = {
                'numero': self.contador,
                'fila': row_num,
                'call_id': call_id,
                'nombre': f"{collected_info.get('firstName', '')} {collected_info.get('lastName', '')}".strip(),
                'telefono': collected_info.get('phoneNumber', ''),
                'clinica': collected_info.get('nombreClinica', ''),
                'flags_json': {
                    'noInteresado': collected_info.get('noInteresado', False),
                    'conPack': collected_info.get('conPack', False),
                    'sinPack': collected_info.get('sinPack', False),
                    'volverALlamar': collected_info.get('volverALlamar', False)
                },
                'nivel_1': nivel_1,
                'nivel_2': nivel_2,
                'payload_valido': payload is not None,
                'tiene_resumen': bool(summary)
            }
            
            self.registros_detallados.append(registro)
            
            # Contar por estado
            estado_key = f"{nivel_1} -> {nivel_2}"
            if estado_key not in self.por_estado:
                self.por_estado[estado_key] = 0
            self.por_estado[estado_key] += 1
            
            # Mostrar progreso cada 10 registros
            if self.contador % 10 == 0:
                print(f"Procesados: {self.contador}")
            
        except Exception as e:
            print(f"ERROR procesando fila {row_num}: {e}")
    
    def mostrar_resumen_final(self):
        """Mostrar resumen detallado final"""
        print()
        print("=" * 80)
        print("RESUMEN FINAL DETALLADO")
        print("=" * 80)
        print(f"Total registros procesados: {self.contador}")
        
        if self.por_estado:
            print(f"\nDistribucion por estado:")
            for estado, cantidad in sorted(self.por_estado.items()):
                porcentaje = (cantidad / self.contador) * 100 if self.contador > 0 else 0
                print(f"  {estado}: {cantidad} ({porcentaje:.1f}%)")
        
        print("\n" + "-" * 60)
        print("DETALLE POR REGISTRO:")
        print("-" * 60)
        print(f"{'#':<3} {'Nombre':<25} {'Estado Final':<35} {'Flags JSON'}")
        print("-" * 60)
        
        for reg in self.registros_detallados:
            flags_str = ""
            if reg['flags_json']['noInteresado']:
                flags_str += "noInt:True "
            if reg['flags_json']['conPack']:
                flags_str += "conPack:True "
            if reg['flags_json']['sinPack']:
                flags_str += "sinPack:True "
            if reg['flags_json']['volverALlamar']:
                flags_str += "volverLlam:True "
            if not flags_str:
                flags_str = "Todos:False"
            
            estado_completo = f"{reg['nivel_1']} -> {reg['nivel_2']}"
            nombre_truncado = reg['nombre'][:24] if len(reg['nombre']) > 24 else reg['nombre']
            
            print(f"{reg['numero']:<3} {nombre_truncado:<25} {estado_completo:<35} {flags_str}")
        
        print("=" * 80)
        
        # Estadisticas adicionales
        payload_validos = sum(1 for r in self.registros_detallados if r['payload_valido'])
        con_resumen = sum(1 for r in self.registros_detallados if r['tiene_resumen'])
        
        print(f"\nESTADISTICAS ADICIONALES:")
        print(f"  Payloads validos: {payload_validos}/{self.contador}")
        print(f"  Registros con resumen: {con_resumen}/{self.contador}")
        print(f"  Registros con flags True: {sum(1 for r in self.registros_detallados if any(r['flags_json'].values()))}")
        
        print("=" * 80)

def main():
    """Funcion principal"""
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = "C:\\Users\\jbeno\\CascadeProjects\\PearlInfo\\SegurcaixaCalls_julio18_procesar_grabaciones.xlsx"
    
    revisor = RevisorCompleto()
    revisor.procesar_archivo_completo(file_path)

if __name__ == "__main__":
    main()