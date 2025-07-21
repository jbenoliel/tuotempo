print("Iniciando script main_segurcaixa.py...")
import os
import requests
import json
import time
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

def load_api_credentials():
    """Cargar credenciales de la API de Pearl desde variables de entorno."""
    load_dotenv()
    account_id = os.getenv('PEARL_ACCOUNT_ID')
    secret_key = os.getenv('PEARL_SECRET_KEY')
    
    if not account_id or not secret_key:
        print("ERROR: No se encontraron las credenciales de la API de Pearl en el archivo .env")
        print("Asegúrate de crear un archivo .env con las variables PEARL_ACCOUNT_ID y PEARL_SECRET_KEY")
        return None, None
    
    return account_id, secret_key

def get_call_info(call_id, account_id, secret_key):
    """Obtener información de una llamada desde la API de Pearl."""
    if not call_id or not account_id or not secret_key:
        return None
    
    url = f"https://api.nlpearl.ai/v1/Call/{call_id}"
    headers = {"Authorization": f"Bearer {account_id}:{secret_key}"}
    
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Error en la API: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"Error al obtener información de la llamada: {e}")
        return None

def extract_json_fields(collected_info_value):
    """Extraer campos específicos del JSON de collected_info."""
    # Inicializar valores por defecto para los campos que vamos a extraer
    fields = {
        "clinicaId": "",
        "certificado": "",
        "codigoPostal": "",
        "delegacion": "",
        "direccionClinica": "",
        "fechaNacimiento": "",
        "firstName": "",
        "lastName": "",
        "nif": "",
        "nombreClinica": "",
        "phoneNumber": "",
        "poliza": "",
        "segmento": "",
        "sexo": ""
    }
    
    if not collected_info_value:
        return fields
    
    try:
        # Intentar decodificar el JSON
        if isinstance(collected_info_value, str):
            # Si es una cadena, intentar parsear como JSON
            data = json.loads(collected_info_value)
        elif isinstance(collected_info_value, dict):
            # Si ya es un diccionario, usarlo directamente
            data = collected_info_value
        else:
            # Si no es ni cadena ni diccionario, no podemos procesarlo
            return fields
        
        # Extraer cada campo si existe en el JSON
        for field in fields.keys():
            if field in data:
                fields[field] = data[field]
        
    except (json.JSONDecodeError, TypeError) as e:
        print(f"Error al decodificar JSON: {e}")
    
    return fields

def process_excel_file(file_path):
    """Leer IDs de llamadas, obtener URLs y procesar el archivo."""
    account_id, secret_key = load_api_credentials()
    if not account_id or not secret_key:
        return

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        header_row = 1

        # Encontrar columnas por encabezado
        call_id_col, recording_col, collected_info_col = None, None, None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                if "Id" in str(cell_value):
                    call_id_col = col
                if "collectedinfo" in str(cell_value).lower():
                    collected_info_col = col
        
        # Forzar columna 2 para Recording URL
        recording_col = 2
        if ws.cell(row=header_row, column=recording_col).value != "Recording URL":
             ws.insert_cols(2)
             ws.cell(row=header_row, column=recording_col, value="Recording URL")

        if not call_id_col:
            print("ADVERTENCIA: No se encontró la columna 'Call ID'")
            return

        # Procesar filas para obtener URLs
        for row in range(2, ws.max_row + 1):
            call_id = ws.cell(row=row, column=call_id_col).value
            if call_id:
                call_info = get_call_info(call_id, account_id, secret_key)
                recording_url = call_info.get("recording") if call_info else None
                ws.cell(row=row, column=recording_col, value=recording_url or "No disponible")

        # Guardar y procesar el archivo con las URLs
        output_file_urls = file_path.replace(".xlsx", "_grabaciones.xlsx")
        wb.save(output_file_urls)
        print(f"\nArchivo con URLs guardado en: {output_file_urls}")
        
        process_and_extract_json_fields(output_file_urls)

    except Exception as e:
        print(f"Error al procesar el archivo Excel: {e}")

def process_and_extract_json_fields(file_path):
    """Procesa un archivo y extrae los campos JSON de la columna CollectedInfo."""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        header_row = 1

        # Encontrar columna CollectedInfo
        collected_info_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value and "collectedinfo" in str(cell_value).lower():
                collected_info_col = col
                print(f"Columna CollectedInfo encontrada: {col}")
                break

        if not collected_info_col:
            print("ADVERTENCIA: No se encontró la columna 'CollectedInfo'")
            return

        # Insertar nuevas columnas para los campos JSON
        current_col = 3  # Empezamos después de Recording URL
        
        # Definir los campos a extraer
        json_fields = [
            "clinicaId", "certificado", "codigoPostal", "delegacion", 
            "direccionClinica", "fechaNacimiento", "firstName", "lastName",
            "nif", "nombreClinica", "phoneNumber", "poliza", "segmento", "sexo"
        ]
        
        # Insertar columnas para cada campo
        field_columns = {}
        for field in json_fields:
            ws.insert_cols(current_col)
            ws.cell(row=header_row, column=current_col, value=field)
            field_columns[field] = current_col
            current_col += 1

        # Procesar filas y extraer campos JSON
        for row in range(2, ws.max_row + 1):
            collected_info = ws.cell(row=row, column=collected_info_col).value
            
            if not collected_info:
                continue

            # Extraer campos del JSON
            fields = extract_json_fields(collected_info)
            
            # Rellenar celdas con los valores extraídos
            for field, value in fields.items():
                col = field_columns.get(field)
                if col:
                    cell = ws.cell(row=row, column=col)
                    cell.value = value

        # Guardar resultados
        # Asegurarse de que no se añada _campos_json a un nombre que ya lo tiene
        if "_campos_json" not in file_path:
            output_file = file_path.replace(".xlsx", "_campos_json.xlsx")
        else:
            output_file = file_path
        wb.save(output_file)
        print(f"Resultados con campos JSON extraídos guardados en: {output_file}")

        # Guardar copia en Dropbox y abrir
        try:
            # Crear nombre de archivo final basado en el nombre original sin la ruta
            nombre_base = os.path.basename(file_path)
            if "_procesar" in nombre_base:
                # Si es un archivo temporal, usar el nombre original
                nombre_base = "julio 21_grabaciones.xlsx"
            else:
                # Si no es temporal, usar el nombre base con el sufijo
                nombre_base = nombre_base.replace(".xlsx", "_grabaciones.xlsx")
                
            dropbox_dir = "C:\\Users\\jbeno\\Dropbox\\TEYAME\\Prueba Segurcaixa"
            dropbox_output_file = os.path.join(dropbox_dir, nombre_base)
            
            print(f"\nGuardando archivo final en Dropbox: {dropbox_output_file}")
            shutil.copy2(output_file, dropbox_output_file)
            print(f"Archivo guardado exitosamente en Dropbox")
            
            # Abrir el archivo
            print("Abriendo archivo final...")
            os.startfile(dropbox_output_file)
            print("Archivo abierto correctamente.")
            
        except Exception as e:
            print(f"\nERROR: No se pudo copiar o abrir el archivo en Dropbox: {e}")
            try:
                print("Intentando abrir archivo local...")
                os.startfile(output_file)
                print("Archivo local abierto correctamente.")
            except Exception as e2:
                print(f"ERROR: No se pudo abrir el archivo automáticamente: {e2}")

    except Exception as e:
        print(f"Error al extraer campos JSON: {e}")

def limpiar_archivos_temporales():
    """Elimina archivos temporales de ejecuciones anteriores."""
    archivos_temp = [
        os.path.join(os.getcwd(), "SegurcaixaCalls_julio21_procesar.xlsx"),
        os.path.join(os.getcwd(), "SegurcaixaCalls_julio21_procesar_grabaciones.xlsx"),
        os.path.join(os.getcwd(), "SegurcaixaCalls_julio21_procesar_grabaciones_campos_json.xlsx")
    ]
    
    for archivo in archivos_temp:
        if os.path.exists(archivo):
            try:
                os.remove(archivo)
                print(f"Archivo temporal eliminado: {archivo}")
            except Exception as e:
                print(f"No se pudo eliminar el archivo temporal {archivo}: {e}")

if __name__ == "__main__":
    print("\n" + "=" * 50)
    print("INICIANDO PROCESAMIENTO DE LLAMADAS SEGURCAIXA")
    print("=" * 50 + "\n")
    
    # Limpiar archivos temporales de ejecuciones anteriores
    limpiar_archivos_temporales()
    
    # Ruta del archivo original en Dropbox
    original_file_path = "C:\\Users\\jbeno\\Dropbox\\TEYAME\\Prueba Segurcaixa\\llamadas julio 21.xlsx"
    local_file_path = os.path.join(os.getcwd(), "SegurcaixaCalls_julio21_procesar.xlsx")
    excel_file = None

    if os.path.exists(original_file_path):
        try:
            print("Copiando archivo original desde Dropbox...")
            original_stats = os.stat(original_file_path)
            shutil.copy2(original_file_path, local_file_path)
            print(f"Archivo original copiado a: {local_file_path}")
            
            # Verificar que la copia se realizó correctamente
            if os.path.exists(local_file_path):
                local_stats = os.stat(local_file_path)
                print(f"Tamaño del archivo local: {local_stats.st_size} bytes")
                print(f"Fecha de modificación local: {time.ctime(local_stats.st_mtime)}")
                if original_stats.st_size == local_stats.st_size:
                    print("Los archivos son idénticos en tamaño.")
                else:
                    print("ADVERTENCIA: Los tamaños de los archivos no coinciden.")
            
            excel_file = local_file_path
        except Exception as e:
            print(f"Error al copiar el archivo desde Dropbox: {e}")
            print("Se intentará usar el archivo original directamente.")
            excel_file = original_file_path
    else:
        print(f"ADVERTENCIA: El archivo original no existe en la ruta: {original_file_path}")
        # Usar el archivo local si existe, o mostrar un error
        if os.path.exists(local_file_path):
            print("Usando archivo local existente.")
            excel_file = local_file_path
        else:
            print("ERROR: No se encontró ni el archivo original ni una copia local. El programa no puede continuar.")
            excel_file = None
    
    if excel_file:
        print(f"\nUsando archivo: {excel_file}")
        try:
            # Registrar tiempo de inicio
            tiempo_inicio = time.time()
            
            # Procesar el archivo
            process_excel_file(excel_file)
            
            # Calcular tiempo de procesamiento
            tiempo_fin = time.time()
            duracion = tiempo_fin - tiempo_inicio
            
            print("\n" + "=" * 50)
            print(f"PROCESAMIENTO COMPLETADO EN {duracion:.2f} SEGUNDOS")
            print("=" * 50)
            print(f"\nArchivo original: {original_file_path}")
            print(f"Archivo con grabaciones: {os.path.join(os.getcwd(), 'SegurcaixaCalls_julio21_procesar_grabaciones.xlsx')}")
            print(f"Archivo final con campos JSON: {os.path.join(os.getcwd(), 'SegurcaixaCalls_julio21_procesar_grabaciones_campos_json.xlsx')}")
            print("\nEl archivo final también se ha guardado en Dropbox.")
            print("=" * 50)
        except Exception as e:
            print(f"\nERROR GENERAL DURANTE EL PROCESAMIENTO: {e}")
            print("\nPor favor, verifica que el archivo original existe y tiene el formato correcto.")
    else:
        print("\nERROR: No se pudo encontrar o crear un archivo Excel para procesar.")
        print("Por favor, asegúrate de que el archivo original existe en la ruta de Dropbox especificada.")
    
    print("\nFin del script.")
    input("\nPresiona Enter para salir...")
