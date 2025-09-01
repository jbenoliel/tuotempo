#!/usr/bin/env python3
"""
Inspeccionar las columnas del archivo Excel para ver si hay URLs de grabación
"""

from openpyxl import load_workbook
import os

def inspeccionar_columnas():
    file_path = r"C:\Users\jbeno\CascadeProjects\tuotempo\Llamadas_07_30_2025_08_01_2025_procesar_grabaciones_campos_json.xlsx"
    
    print(f"Inspeccionando columnas de: {os.path.basename(file_path)}")
    print("="*60)
    
    if not os.path.exists(file_path):
        print(f"ERROR: Archivo no encontrado: {file_path}")
        return
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print("COLUMNAS ENCONTRADAS:")
        print("-" * 40)
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                print(f"Columna {col:2d}: {cell_value}")
                
                # Buscar específicamente columnas relacionadas con grabaciones
                cell_lower = str(cell_value).lower()
                if any(keyword in cell_lower for keyword in ['recording', 'grabacion', 'grabación', 'audio', 'url', 'link']):
                    print(f"    *** POSIBLE COLUMNA DE GRABACIÓN ***")
        
        print("\n" + "="*60)
        print(f"Total de columnas: {ws.max_column}")
        print(f"Total de filas: {ws.max_row}")
        
        # Mostrar una muestra de la primera fila de datos
        print("\nMUESTRA DE PRIMERA FILA DE DATOS:")
        print("-" * 40)
        for col in range(1, min(ws.max_column + 1, 6)):  # Mostrar solo las primeras 5 columnas
            header = ws.cell(row=1, column=col).value
            data = ws.cell(row=2, column=col).value
            print(f"{header}: {str(data)[:50]}{'...' if len(str(data)) > 50 else ''}")
        
    except Exception as e:
        print(f"ERROR: {e}")

if __name__ == "__main__":
    inspeccionar_columnas()
