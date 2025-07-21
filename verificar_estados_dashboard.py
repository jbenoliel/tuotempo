#!/usr/bin/env python3
"""
Verificar estados específicos en el dashboard vs nuestros datos
"""
import os
import json
from openpyxl import load_workbook

def analizar_discrepancia():
    """Analizar la discrepancia entre Excel y Dashboard"""
    file_path = "C:\\Users\\jbeno\\CascadeProjects\\PearlInfo\\SegurcaixaCalls_julio18_procesar_grabaciones.xlsx"
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Encontrar columna CollectedInfo
    collected_info_col = None
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value and "collectedinfo" in str(cell_value).lower():
            collected_info_col = col
            break
    
    print("=== DISCREPANCIA EXCEL vs DASHBOARD ===")
    print()
    print("EXCEL dice:")
    print("- 5 registros con noInteresado: true")
    print("- 34 registros sin flags (procesados como 'Volver a llamar')")
    print()
    print("DASHBOARD dice:")
    print("- 8 registros 'No Interesado' (6 con subestado)")
    print("- 40 registros 'Volver a llamar' (con subestados)")
    print()
    print("CONCLUSIÓN:")
    print("- Hay 3 registros adicionales marcados como 'No Interesado' en el dashboard")
    print("- Estos 3 registros NO tienen noInteresado: true en el Excel")
    print("- Fueron marcados manualmente o por otro proceso")
    print()
    print("REGISTROS CON noInteresado: true EN EXCEL:")
    
    # Mostrar los 5 registros del Excel
    registros_excel = []
    for row in range(2, ws.max_row + 1):
        collected_info_raw = ws.cell(row=row, column=collected_info_col).value
        
        try:
            if isinstance(collected_info_raw, str):
                data = json.loads(collected_info_raw)
            elif isinstance(collected_info_raw, dict):
                data = collected_info_raw
            else:
                continue
            
            if data.get('noInteresado', False):
                nombre = f"{data.get('firstName', '')} {data.get('lastName', '')}".strip()
                telefono = data.get('phoneNumber', '')
                # Limpiar teléfono
                if telefono.startswith('+34'):
                    telefono = telefono[3:]
                elif telefono.startswith('34') and len(telefono) == 11:
                    telefono = telefono[2:]
                    
                registros_excel.append({
                    'fila': row,
                    'nombre': nombre,
                    'telefono': telefono
                })
        except:
            continue
    
    for i, reg in enumerate(registros_excel, 1):
        print(f"{i}. {reg['nombre']} ({reg['telefono']})")
    
    print()
    print("RECOMENDACIÓN:")
    print("1. Verifica en el dashboard cuáles son exactamente los 8 registros 'No Interesado'")
    print("2. Identifica los 3 que NO están en la lista anterior")
    print("3. Esos 3 registros necesitan ser corregidos manualmente en el dashboard")
    print("   O puedes cambiar su estado a 'Volver a llamar' si es incorrecto")
    print()
    print("PARA CORREGIR LOS 2 SUBESTADOS FALTANTES:")
    print("- Necesitas identificar cuáles de los 8 'No Interesado' del dashboard")
    print("- NO tienen subestado")
    print("- Y aplicarles el subestado 'No da motivos' manualmente")

if __name__ == "__main__":
    analizar_discrepancia()