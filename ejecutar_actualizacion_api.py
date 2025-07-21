"""
Ejecutor de Actualización API - Llamadas Reales

Este programa ejecuta las llamadas REALES a la API de actualizar_resultado
usando los datos del JSON collectedInfo.
"""

import os
import json
import time
import logging
import requests
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
from mapeo_inteligente_segurcaixa import MapeadorInteligente

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('actualizacion_api_real.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EjecutorAPI:
    def __init__(self):
        load_dotenv()
        self.api_base_url = "https://tuotempo-apis-production.up.railway.app"
        self.mapeador = MapeadorInteligente()
        self.estadisticas = {
            'total_procesados': 0,
            'exitosos': 0,
            'errores_api': 0,
            'errores_validacion_bd': 0,
            'sin_telefono': 0,
            'estados_aplicados': {}
        }
        
        logger.info("=== EJECUTOR API REAL INICIADO ===")
    
    def extract_json_fields(self, collected_info_value):
        """Extraer campos del JSON collectedInfo"""
        fields = {
            # Campos de cliente
            "clinicaId": "", "certificado": "", "codigoPostal": "", "delegacion": "",
            "direccionClinica": "", "fechaNacimiento": "", "firstName": "", "lastName": "",
            "nif": "", "nombreClinica": "", "phoneNumber": "", "poliza": "", "segmento": "", "sexo": "",
            # Campos de resultado de llamada
            "noInteresado": False, "volverALlamar": False, "conPack": False, "sinPack": False,
            "fechaEscogida": "", "horaEscogida": "", "razonNoInteres": "", "razonVolverALlamar": ""
        }
        
        if not collected_info_value:
            return fields
        
        try:
            if isinstance(collected_info_value, str):
                data = json.loads(collected_info_value)
            elif isinstance(collected_info_value, dict):
                data = collected_info_value
            else:
                return fields
            
            for field in fields.keys():
                if field in data:
                    fields[field] = data[field]
        except:
            pass
        
        return fields
    
    def call_api(self, payload):
        """Realizar llamada REAL a la API actualizar_resultado"""
        url = f"{self.api_base_url}/api/actualizar_resultado"
        headers = {'Content-Type': 'application/json'}
        
        try:
            logger.info(f"Llamando API para teléfono {payload['telefono']}")
            response = requests.post(url, json=payload, headers=headers, timeout=30)
            
            if response.status_code == 200:
                result = response.json()
                logger.info(f"API ÉXITO {payload['telefono']}: {result.get('message', 'OK')}")
                return True, result, response.status_code
            else:
                logger.error(f"API ERROR {payload['telefono']}: {response.status_code} - {response.text}")
                return False, {"error": response.text}, response.status_code
                
        except requests.exceptions.RequestException as e:
            logger.error(f"CONEXIÓN ERROR {payload['telefono']}: {e}")
            return False, {"error": str(e)}, 0
    
    def verificar_en_bd(self, telefono):
        """Verificación simplificada - confiar en respuesta de API actualizar_resultado"""
        # Por ahora, no hacemos verificación adicional ya que el endpoint consultar_lead no existe
        # La verificación se basa en el resultado de la API actualizar_resultado
        logger.info(f"BD VERIFICACIÓN: Confiando en respuesta de API actualizar_resultado para {telefono}")
        return {"status": "API_SUCCESS", "telefono": telefono}
    
    def ejecutar_actualizacion(self, file_path):
        """Ejecutar actualización completa del archivo"""
        logger.info(f"=== INICIANDO ACTUALIZACIÓN REAL ===")
        logger.info(f"Archivo: {file_path}")
        logger.info(f"API: {self.api_base_url}")
        
        if not os.path.exists(file_path):
            logger.error(f"Archivo no encontrado: {file_path}")
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Encontrar columnas
            collected_info_col = None
            call_id_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    cell_lower = str(cell_value).lower()
                    if "collectedinfo" in cell_lower:
                        collected_info_col = col
                    elif "id" in cell_lower and call_id_col is None:
                        call_id_col = col
            
            if not collected_info_col:
                logger.error("ERROR: No se encontró la columna 'CollectedInfo'")
                return
            
            total_filas = ws.max_row - 1
            logger.info(f"Total registros a procesar: {total_filas}")
            
            # Procesar cada fila
            for row in range(2, ws.max_row + 1):
                try:
                    self.procesar_fila_real(ws, row, collected_info_col, call_id_col)
                    
                    # Pausa entre llamadas para no saturar la API
                    time.sleep(1)
                    
                    # Mostrar progreso cada 10 registros
                    if self.estadisticas['total_procesados'] % 10 == 0:
                        logger.info(f"Procesados: {self.estadisticas['total_procesados']}/{total_filas}")
                    
                except Exception as e:
                    logger.error(f"Error procesando fila {row}: {e}")
            
            self.mostrar_resumen_final()
            
        except Exception as e:
            logger.error(f"Error al procesar archivo: {e}")
    
    def procesar_fila_real(self, worksheet, row_num, collected_info_col, call_id_col):
        """Procesar una fila individual con llamada REAL a la API"""
        try:
            self.estadisticas['total_procesados'] += 1
            
            # Extraer datos
            collected_info_raw = worksheet.cell(row=row_num, column=collected_info_col).value
            call_id = worksheet.cell(row=row_num, column=call_id_col).value if call_id_col else ""
            
            # Extraer campos del JSON
            collected_info = self.extract_json_fields(collected_info_raw)
            
            nombre = f"{collected_info.get('firstName', '')} {collected_info.get('lastName', '')}".strip()
            telefono = collected_info.get('phoneNumber', '')
            
            logger.info(f"=== PROCESANDO {self.estadisticas['total_procesados']}: {nombre} ({telefono}) ===")
            
            # Generar payload
            payload = self.mapeador.mapear_a_api_payload_desde_json(collected_info)
            
            if not payload:
                logger.warning(f"No se pudo crear payload para {nombre}")
                self.estadisticas['sin_telefono'] += 1
                return
            
            # Determinar estado para estadísticas
            if payload.get('noInteresado'):
                estado = 'No Interesado'
            elif payload.get('conPack'):
                estado = 'CONFIRMADO Con Pack'
            elif payload.get('nuevaCita'):
                estado = 'CONFIRMADO Sin Pack'
            elif payload.get('volverALlamar'):
                estado = 'Volver a llamar'
            else:
                estado = 'Sin clasificar'
            
            logger.info(f"Estado detectado: {estado}")
            logger.info(f"Payload: {payload}")
            
            # Llamar a la API REAL
            success, response, status_code = self.call_api(payload)
            
            if success:
                self.estadisticas['exitosos'] += 1
                
                # Contar estados aplicados
                if estado not in self.estadisticas['estados_aplicados']:
                    self.estadisticas['estados_aplicados'][estado] = 0
                self.estadisticas['estados_aplicados'][estado] += 1
                
                # Verificar en BD
                time.sleep(0.5)  # Esperar un poco para que se aplique el cambio
                bd_result = self.verificar_en_bd(payload['telefono'])
                
                if bd_result:
                    logger.info(f"EXITO COMPLETO: {nombre}")
                else:
                    logger.warning(f"API OK pero no se verifico en BD: {nombre}")
                    self.estadisticas['errores_validacion_bd'] += 1
            else:
                logger.error(f"ERROR API: {nombre}")
                self.estadisticas['errores_api'] += 1
            
        except Exception as e:
            logger.error(f"Error procesando fila {row_num}: {e}")
            self.estadisticas['errores_api'] += 1
    
    def mostrar_resumen_final(self):
        """Mostrar resumen final de la ejecución"""
        logger.info("\n" + "="*80)
        logger.info("=== RESUMEN FINAL DE ACTUALIZACIÓN API ===")
        logger.info("="*80)
        logger.info(f"Total procesados: {self.estadisticas['total_procesados']}")
        logger.info(f"Exitosos: {self.estadisticas['exitosos']}")
        logger.info(f"Errores API: {self.estadisticas['errores_api']}")
        logger.info(f"Errores validación BD: {self.estadisticas['errores_validacion_bd']}")
        logger.info(f"Sin teléfono: {self.estadisticas['sin_telefono']}")
        
        if self.estadisticas['total_procesados'] > 0:
            tasa_exito = (self.estadisticas['exitosos'] / self.estadisticas['total_procesados']) * 100
            logger.info(f"Tasa de éxito: {tasa_exito:.2f}%")
        
        if self.estadisticas['estados_aplicados']:
            logger.info("\nEstados aplicados:")
            for estado, cantidad in sorted(self.estadisticas['estados_aplicados'].items()):
                logger.info(f"  {estado}: {cantidad}")
        
        logger.info("="*80)

def main():
    """Función principal"""
    import sys
    
    print("\n" + "="*80)
    print("EJECUTOR DE ACTUALIZACION API REAL")
    print("="*80)
    print("ATENCION: Este programa realizara llamadas REALES a la API")
    print("y modificara los datos en la base de datos de produccion.")
    print("="*80)
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        print(f"Archivo especificado: {file_path}")
    else:
        file_path = input("\nRuta del archivo Excel: ").strip()
        if not file_path:
            file_path = "C:\\Users\\jbeno\\CascadeProjects\\PearlInfo\\SegurcaixaCalls_julio18_procesar_grabaciones.xlsx"
    
    confirm = input(f"\nCONFIRMACION FINAL\nEjecutar actualizacion REAL para {os.path.basename(file_path)}?\nEsto modificara la base de datos de produccion.\n(escriba 'CONFIRMO' para continuar): ")
    
    if confirm != 'CONFIRMO':
        print("Operacion cancelada.")
        return
    
    ejecutor = EjecutorAPI()
    ejecutor.ejecutar_actualizacion(file_path)
    
    print(f"\nEjecucion completada. Ver log: actualizacion_api_real.log")

if __name__ == "__main__":
    main()